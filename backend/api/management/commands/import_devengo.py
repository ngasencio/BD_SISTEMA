"""
Comando de gestión Django para importar datos del XLSX de devengo.

Uso:
    python manage.py import_devengo --file "ruta/al/bd_devengo.xlsx"
"""
import openpyxl
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
from django.core.management.base import BaseCommand
from api.models import Devengo


class Command(BaseCommand):
    help = 'Importa registros de devengo desde un archivo XLSX'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Ruta al archivo XLSX de devengo'
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Limpiar tabla antes de importar'
        )

    def handle(self, *args, **options):
        filepath = options['file']

        if options['clear']:
            count = Devengo.objects.count()
            Devengo.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'Tabla limpiada: {count} registros eliminados'))

        self.stdout.write(f'Abriendo archivo: {filepath}')

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            self.stderr.write(self.style.ERROR(f'Error al abrir XLSX: {e}'))
            return

        # Leer encabezados de la primera fila
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value).strip() if cell.value else '')

        self.stdout.write(f'Columnas detectadas: {len(headers)}')

        # Mapeo de columnas del XLSX a campos del modelo
        col_map = {
            'Código Unidad Ejecutora': 'codigo_ue',
            'Folio': 'folio',
            'Titulo': 'titulo',
            'Tipo Presupuesto': 'tipo_presupuesto',
            'Moneda Presupuestaria': 'moneda_presupuestaria',
            'Principal': 'principal',
            'Principal Relacionado': 'principal_relacionado',
            'Moneda Documento': 'moneda_documento',
            'Tipo Cambio': 'tipo_cambio',
            'Tipo Documento': 'tipo_documento',
            'Número Documento': 'numero_documento',
            'Fecha Documento': 'fecha_documento',
            'Fecha Conforme': 'fecha_conforme',
            'Id Chile Compra': 'id_chile_compra',
            'Fecha Ingreso / Fecha Recepción ': 'fecha_ingreso',
            'Fecha Ingreso': 'fecha_ingreso',
            'Fecha Recepción': 'fecha_ingreso',
            'Catálogo 01': 'catalogo_01',
            'Catálogo 02': 'catalogo_02',
            'Catálogo 03': 'catalogo_03',
            'Catálogo 04': 'catalogo_04',
            'Catálogo 05': 'catalogo_05',
            'Catálogos 01': 'catalogo_01',
            'Catálogos 02': 'catalogo_02',
            'Catálogos 03': 'catalogo_03',
            'Catálogos 04': 'catalogo_04',
            'Catálogos 05': 'catalogo_05',
            'Concepto Presupuestario': 'concepto_presupuestario',
            'Monto Vigente': 'monto_vigente',
            'Monto Disponible': 'monto_disponible',
            'Monto Consumido': 'monto_consumido',
            'archivo_origen': 'archivo_origen',
        }

        # Índices de columnas según encabezado
        col_idx = {}
        for i, h in enumerate(headers):
            if h in col_map:
                col_idx[col_map[h]] = i

        self.stdout.write(f'Campos mapeados: {list(col_idx.keys())}')

        imported = 0
        errors = 0
        batch = []
        BATCH_SIZE = 500

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                obj = Devengo()

                for field, idx in col_idx.items():
                    if idx >= len(row):
                        continue
                    val = row[idx]

                    if val is None or val == '':
                        continue

                    # Campos de fecha
                    if field in ('fecha_documento', 'fecha_conforme', 'fecha_ingreso'):
                        if isinstance(val, (datetime, date)):
                            setattr(obj, field, val if isinstance(val, date) else val.date())
                        elif isinstance(val, str):
                            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
                                try:
                                    setattr(obj, field, datetime.strptime(val.strip(), fmt).date())
                                    break
                                except ValueError:
                                    pass

                    # Campos decimales
                    elif field in ('monto_vigente', 'monto_disponible', 'monto_consumido', 'tipo_cambio'):
                        try:
                            setattr(obj, field, Decimal(str(val)).quantize(Decimal('0.01')))
                        except (InvalidOperation, ValueError):
                            setattr(obj, field, Decimal('0'))

                    # Campos de texto
                    else:
                        setattr(obj, field, str(val).strip()[:500] if val else None)

                # Asegurar campo obligatorio
                if not obj.codigo_ue:
                    continue

                batch.append(obj)

                if len(batch) >= BATCH_SIZE:
                    Devengo.objects.bulk_create(batch, ignore_conflicts=True)
                    imported += len(batch)
                    self.stdout.write(f'  → {imported} registros importados...')
                    batch = []

            except Exception as e:
                errors += 1
                if errors <= 10:
                    self.stderr.write(f'  Error fila {row_num}: {e}')

        # Importar el último lote
        if batch:
            Devengo.objects.bulk_create(batch, ignore_conflicts=True)
            imported += len(batch)

        wb.close()

        self.stdout.write(self.style.SUCCESS(
            f'\n✅ Importación completada: {imported} registros importados, {errors} errores.'
        ))
